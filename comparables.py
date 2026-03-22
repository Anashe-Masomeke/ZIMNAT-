import streamlit as st
import pandas as pd
import numpy as np
import yfinance as yf
from pathlib import Path
import re
from bs4 import BeautifulSoup
from io import StringIO
import requests
import time
import random

st.set_page_config(page_title="Comparables Valuation (Excel Style)", layout="wide")
st.title("📊 Comparables Valuation – EV/EBITDA, P/B, P/E")
st.markdown("All values + inputs are **saved in session_state**, so switching pages keeps your work.")

S = st.session_state


# =========================================================
# HELPERS
# =========================================================
def format_numeric_columns(df):
    fmt = {}
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            fmt[col] = "{:,.2f}"
    return df.style.format(fmt)


def _clean_text(x) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    return str(x).strip()


def _norm_text(x: str) -> str:
    return re.sub(r"[^a-z0-9]", "", _clean_text(x).lower())


def _tokenize_text(x: str) -> list:
    return [t for t in re.split(r"[^a-z0-9]+", _clean_text(x).lower()) if t]


def _clean_num(x):
    try:
        if x is None or x == "":
            return np.nan
        return float(x)
    except Exception:
        return np.nan


def _num_input_default(x, fallback=0.0):
    try:
        if x is None or pd.isna(x):
            return float(fallback)
        return float(x)
    except Exception:
        return float(fallback)


def normalize_peer_ticker(symbol: str) -> str:
    sym = _clean_text(symbol).upper()
    fixes = {
        "MTNN": "MTNN.NG",
        "SCOM": "SCOM.KE",
        "SAFARICOM": "SCOM.KE",
        "EQTY": "EQTY.KE",
        "KCB": "KCB.KE",
        "VODACOM": "VOD.JO",
        "MTN": "MTN.JO",
    }
    return fixes.get(sym, sym)


def make_yahoo_profile_url(symbol: str) -> str:
    sym = normalize_peer_ticker(symbol)
    return f"https://finance.yahoo.com/quote/{sym}/profile/" if sym else ""


def filtered_average(values, band=0.4):
    arr = np.array(values, dtype=float)
    arr = arr[~np.isnan(arr)]
    arr = arr[arr != 0]

    if len(arr) == 0:
        return np.nan

    median = np.median(arr)
    lower = median * (1 - band)
    upper = median * (1 + band)
    keep = arr[(arr >= lower) & (arr <= upper)]
    return float(np.mean(keep if len(keep) > 0 else arr))

# =========================================================
# UNIVERSE FILE
# =========================================================
def find_universe_file() -> str:
    candidates = [
        "data/africa_yahoo_peer_universe_strict_final.xlsx",
        "data/africa_yahoo_peer_universe_starter.xlsx",
        "africa_yahoo_peer_universe_strict_final.xlsx",
        "africa_yahoo_peer_universe_starter.xlsx",
        "/mnt/data/africa_yahoo_peer_universe_strict_final.xlsx",
        "/mnt/data/africa_yahoo_peer_universe_starter.xlsx",
    ]
    for p in candidates:
        if Path(p).exists():
            return p
    return ""


def _std_colname(x: str) -> str:
    x = _clean_text(x).lower()
    x = re.sub(r"[^a-z0-9]+", "_", x)
    return x.strip("_")


def _find_sheet_name(xls: pd.ExcelFile, wanted_names: list, fallback_contains: list = None):
    sheets = list(xls.sheet_names)
    norm_map = {_std_colname(s): s for s in sheets}

    for w in wanted_names:
        wn = _std_colname(w)
        if wn in norm_map:
            return norm_map[wn]

    if fallback_contains:
        for s in sheets:
            sn = _std_colname(s)
            if all(tok in sn for tok in fallback_contains):
                return s

    return None


@st.cache_data(show_spinner=False)
def load_peer_universe_bundle(path: str):
    if not path or not Path(path).exists():
        return None, None, None, {"error": "Universe file not found."}

    try:
        xls = pd.ExcelFile(path)
    except Exception as e:
        return None, None, None, {"error": f"Could not open Excel file: {e}"}

    universe_sheet = _find_sheet_name(
        xls,
        wanted_names=["CODE_READY_UNIVERSE", "code_ready_universe", "UNIVERSE", "peer_universe"],
        fallback_contains=["universe"]
    )
    zim_map_sheet = _find_sheet_name(
        xls,
        wanted_names=["ZIM_TARGET_MAP", "zim_target_map", "TARGET_MAP"],
        fallback_contains=["target", "map"]
    )
    alias_sheet = _find_sheet_name(
        xls,
        wanted_names=["SECTOR_ALIAS_MAP", "sector_alias_map", "ALIAS_MAP"],
        fallback_contains=["alias"]
    )

    debug = {
        "file": path,
        "sheet_names": list(xls.sheet_names),
        "universe_sheet": universe_sheet,
        "zim_map_sheet": zim_map_sheet,
        "alias_sheet": alias_sheet,
    }

    if universe_sheet is None:
        return None, None, None, {
            **debug,
            "error": "Could not find the universe sheet in the Excel file."
        }

    universe = pd.read_excel(xls, sheet_name=universe_sheet)
    zim_map = pd.read_excel(xls, sheet_name=zim_map_sheet) if zim_map_sheet else pd.DataFrame()
    alias_map = pd.read_excel(xls, sheet_name=alias_sheet) if alias_sheet else pd.DataFrame()

    universe.columns = [_std_colname(c) for c in universe.columns]
    zim_map.columns = [_std_colname(c) for c in zim_map.columns]
    alias_map.columns = [_std_colname(c) for c in alias_map.columns]

    # expected universe column aliases
    rename_universe = {
        "symbol": "ticker",
        "peer_ticker": "ticker",
        "peer_symbol": "ticker",
        "company_name": "company",
        "peer_company": "company",
        "peer_name": "company",
        "country_name": "country",
        "sector_name": "sector",
        "industry_name": "industry",
        "keywords": "sector_keywords",
        "peer_keywords": "sector_keywords",
        "priority": "match_priority",
        "yahoo_confirmed": "yahoo_status",
        "status": "yahoo_status",
        "exchange_name": "exchange",
    }
    universe = universe.rename(columns={k: v for k, v in rename_universe.items() if k in universe.columns})

    rename_zim = {
        "symbol": "target_symbol",
        "company": "target_company",
        "sector": "preferred_sector",
        "industry": "preferred_industry",
        "keywords": "preferred_peer_keywords",
    }
    zim_map = zim_map.rename(columns={k: v for k, v in rename_zim.items() if k in zim_map.columns})

    rename_alias = {
        "alias": "input_alias",
        "sector": "preferred_sector",
        "industry": "preferred_industry",
    }
    alias_map = alias_map.rename(columns={k: v for k, v in rename_alias.items() if k in alias_map.columns})

    # ensure required columns exist
    for col in ["ticker", "company", "country", "exchange", "sector", "industry", "sector_keywords", "match_priority", "yahoo_status"]:
        if col not in universe.columns:
            universe[col] = ""

    for col in ["target_symbol", "target_company", "preferred_sector", "preferred_industry", "preferred_peer_keywords", "search_aliases"]:
        if col not in zim_map.columns:
            zim_map[col] = ""

    for col in ["input_alias", "preferred_sector", "preferred_industry"]:
        if col not in alias_map.columns:
            alias_map[col] = ""

    # clean all values
    for df in [universe, zim_map, alias_map]:
        for col in df.columns:
            df[col] = df[col].map(_clean_text)

    universe["ticker"] = universe["ticker"].map(normalize_peer_ticker)
    zim_map["target_symbol"] = zim_map["target_symbol"].map(lambda x: _clean_text(x).upper())
    alias_map["input_alias"] = alias_map["input_alias"].map(lambda x: _clean_text(x).lower())

    # remove blank tickers
    universe = universe[universe["ticker"].map(lambda x: _clean_text(x) != "")].copy()

    debug["universe_columns"] = list(universe.columns)
    debug["zim_map_columns"] = list(zim_map.columns)
    debug["alias_map_columns"] = list(alias_map.columns)
    debug["universe_rows"] = len(universe)
    debug["zim_map_rows"] = len(zim_map)
    debug["alias_map_rows"] = len(alias_map)

    return universe, zim_map, alias_map, debug


UNIVERSE_FILE = find_universe_file()
UNIVERSE_DF, ZIM_TARGET_MAP_DF, SECTOR_ALIAS_DF, UNIVERSE_DEBUG = load_peer_universe_bundle(UNIVERSE_FILE)
# =========================================================
# FALLBACK ZIM TARGET MAP
# =========================================================
FALLBACK_ZIM_TARGETS = [
    {"target_symbol": "ECOZIM", "target_company": "Econet Wireless Zimbabwe", "preferred_sector": "Telecommunications", "preferred_industry": "Mobile Telecoms", "preferred_peer_keywords": "telecommunications,telecom,mobile,wireless,communications,network,broadband,data"},
    {"target_symbol": "CBZ", "target_company": "CBZ", "preferred_sector": "Banking", "preferred_industry": "Commercial Banks", "preferred_peer_keywords": "banking,bank,commercial bank,lending,deposits"},
    {"target_symbol": "FBC", "target_company": "FBC", "preferred_sector": "Banking", "preferred_industry": "Commercial Banks", "preferred_peer_keywords": "banking,bank,commercial bank,lending,deposits"},
    {"target_symbol": "NMBZ", "target_company": "NMBZ", "preferred_sector": "Banking", "preferred_industry": "Commercial Banks", "preferred_peer_keywords": "banking,bank,commercial bank,lending,deposits"},
    {"target_symbol": "ZB", "target_company": "ZB", "preferred_sector": "Banking", "preferred_industry": "Commercial Banks", "preferred_peer_keywords": "banking,bank,commercial bank,lending,deposits"},
    {"target_symbol": "PADENGA", "target_company": "Padenga", "preferred_sector": "Mining", "preferred_industry": "Gold Mining", "preferred_peer_keywords": "gold mining,gold,gold producer,gold miner,mining,minerals,precious metals"},
    {"target_symbol": "CMCL", "target_company": "Caledonia", "preferred_sector": "Mining", "preferred_industry": "Gold Mining", "preferred_peer_keywords": "gold mining,gold,gold producer,gold miner,mining,minerals,precious metals"},
    {"target_symbol": "DELTA", "target_company": "Delta", "preferred_sector": "Consumer Staples", "preferred_industry": "Beverages", "preferred_peer_keywords": "beverages,brewery,beer,spirits,distillery,drinks"},
    {"target_symbol": "AFDIS", "target_company": "Afdis", "preferred_sector": "Consumer Staples", "preferred_industry": "Beverages", "preferred_peer_keywords": "beverages,brewery,beer,spirits,distillery,drinks"},
    {"target_symbol": "INNSCOR", "target_company": "Innscor Africa", "preferred_sector": "Consumer Staples", "preferred_industry": "Food Producers", "preferred_peer_keywords": "food,consumer,food processing,packaged foods,brands"},
    {"target_symbol": "SIMBISA", "target_company": "Simbisa", "preferred_sector": "Consumer Discretionary", "preferred_industry": "Restaurants", "preferred_peer_keywords": "restaurants,quick service restaurants,foodservice,fast food"},
    {"target_symbol": "WESTPROP", "target_company": "WestProp", "preferred_sector": "Real Estate", "preferred_industry": "Property", "preferred_peer_keywords": "real estate,property,reit,property development"},
    {"target_symbol": "RTG", "target_company": "Rainbow Tourism Group", "preferred_sector": "Consumer Discretionary", "preferred_industry": "Hotels", "preferred_peer_keywords": "hotels,lodging,leisure,tourism"},
    {"target_symbol": "ASUN", "target_company": "African Sun", "preferred_sector": "Consumer Discretionary", "preferred_industry": "Hotels", "preferred_peer_keywords": "hotels,lodging,leisure,tourism"},
]

if ZIM_TARGET_MAP_DF is None or ZIM_TARGET_MAP_DF.empty:
    ZIM_TARGET_MAP_DF = pd.DataFrame(FALLBACK_ZIM_TARGETS)

if SECTOR_ALIAS_DF is None or SECTOR_ALIAS_DF.empty:
    SECTOR_ALIAS_DF = pd.DataFrame([
        {"input_alias": "telecom", "preferred_sector": "Telecommunications", "preferred_industry": "Mobile Telecoms"},
        {"input_alias": "telecommunications", "preferred_sector": "Telecommunications", "preferred_industry": "Mobile Telecoms"},
        {"input_alias": "communication services", "preferred_sector": "Telecommunications", "preferred_industry": "Mobile Telecoms"},
        {"input_alias": "mobile telecoms", "preferred_sector": "Telecommunications", "preferred_industry": "Mobile Telecoms"},
        {"input_alias": "banking", "preferred_sector": "Banking", "preferred_industry": "Commercial Banks"},
        {"input_alias": "banks", "preferred_sector": "Banking", "preferred_industry": "Commercial Banks"},
        {"input_alias": "gold mining", "preferred_sector": "Mining", "preferred_industry": "Gold Mining"},
        {"input_alias": "mining", "preferred_sector": "Mining", "preferred_industry": ""},
        {"input_alias": "beverages", "preferred_sector": "Consumer Staples", "preferred_industry": "Beverages"},
        {"input_alias": "real estate", "preferred_sector": "Real Estate", "preferred_industry": "Property"},
    ])
# =========================================================
# TARGET RESOLUTION
# =========================================================
def find_target_row(query: str):
    q = _clean_text(query)
    qn = _norm_text(q)
    if not qn or ZIM_TARGET_MAP_DF is None or ZIM_TARGET_MAP_DF.empty:
        return None

    for _, r in ZIM_TARGET_MAP_DF.iterrows():
        if _norm_text(r.get("target_symbol", "")) == qn:
            return r.to_dict()

    for _, r in ZIM_TARGET_MAP_DF.iterrows():
        if _norm_text(r.get("target_company", "")) == qn:
            return r.to_dict()

    for _, r in ZIM_TARGET_MAP_DF.iterrows():
        if qn and qn in _norm_text(r.get("target_company", "")):
            return r.to_dict()

    return None


def normalize_sector_override(sector_text: str):
    s = _clean_text(sector_text).lower()
    if not s:
        return "", ""

    for _, r in SECTOR_ALIAS_DF.iterrows():
        if s == _clean_text(r.get("input_alias")).lower():
            return _clean_text(r.get("preferred_sector")), _clean_text(r.get("preferred_industry"))

    return sector_text.strip(), ""


def split_keywords(text: str):
    out = []
    for x in _clean_text(text).split(","):
        x = _clean_text(x).lower()
        if x:
            out.append(x)
    return out


def get_target_profile(target_query: str, manual_sector_override: str = ""):
    row = find_target_row(target_query)

    if row:
        target_symbol = _clean_text(row.get("target_symbol")).upper()
        target_company = _clean_text(row.get("target_company"))
        preferred_sector = _clean_text(row.get("preferred_sector"))
        preferred_industry = _clean_text(row.get("preferred_industry"))
        preferred_keywords = split_keywords(row.get("preferred_peer_keywords"))
        search_aliases = split_keywords(row.get("search_aliases", ""))
        source = "ZIM_TARGET_MAP"
    else:
        target_symbol = _clean_text(target_query).upper()
        target_company = _clean_text(target_query)
        preferred_sector = ""
        preferred_industry = ""
        preferred_keywords = []
        search_aliases = []
        source = "manual"

    if manual_sector_override.strip():
        sec, ind = normalize_sector_override(manual_sector_override)
        preferred_sector = sec or preferred_sector
        preferred_industry = ind
        source = f"{source} + manual_sector_override"
    if not preferred_keywords:
        preferred_keywords = split_keywords(f"{preferred_sector},{preferred_industry}")

    return {
        "target_symbol": target_symbol,
        "target_company": target_company,
        "preferred_sector": preferred_sector,
        "preferred_industry": preferred_industry,
        "preferred_peer_keywords": preferred_keywords,
        "search_aliases": search_aliases,
        "source": source,
    }
# =========================================================
# UNIVERSE FILTERING
# =========================================================
def strict_peer_score(peer_row: dict, target_profile: dict) -> int:
    score = 0

    peer_sector = _clean_text(peer_row.get("sector")).lower()
    peer_industry = _clean_text(peer_row.get("industry")).lower()
    peer_keywords = _clean_text(peer_row.get("sector_keywords")).lower()
    peer_company = _clean_text(peer_row.get("company")).lower()
    peer_country = _clean_text(peer_row.get("country")).lower()
    peer_priority = _clean_num(peer_row.get("match_priority"))

    tgt_sector = _clean_text(target_profile.get("preferred_sector")).lower()
    tgt_industry = _clean_text(target_profile.get("preferred_industry")).lower()
    tgt_keywords = [k.lower() for k in target_profile.get("preferred_peer_keywords", [])]

    combo = f"{peer_sector} | {peer_industry} | {peer_keywords} | {peer_company}"

    if tgt_sector and peer_sector == tgt_sector:
        score += 50
    elif tgt_sector and (tgt_sector in peer_sector or peer_sector in tgt_sector):
        score += 35
    elif tgt_sector and tgt_sector in combo:
        score += 20

    if tgt_industry and peer_industry == tgt_industry:
        score += 60
    elif tgt_industry and (tgt_industry in peer_industry or peer_industry in tgt_industry):
        score += 40
    elif tgt_industry and tgt_industry in combo:
        score += 20

    for kw in tgt_keywords:
        if kw and kw in peer_industry:
            score += 15
        if kw and kw in peer_keywords:
            score += 12
        if kw and kw in peer_sector:
            score += 10
        if kw and kw in peer_company:
            score += 4

    if not pd.isna(peer_priority):
        score += int(peer_priority) * 5

    if peer_country in ["south africa", "kenya", "nigeria", "zimbabwe", "botswana", "egypt", "ghana", "mauritius"]:
        score += 3

    return score


def _family_words(target_profile: dict):
    tgt_sector = _clean_text(target_profile.get("preferred_sector")).lower()
    tgt_industry = _clean_text(target_profile.get("preferred_industry")).lower()

    if tgt_sector == "telecommunications" or tgt_industry in ["mobile telecoms", "telecoms", "telecommunications"]:
        return {
            "good": [
                "telecom", "telecommunications", "communication services",
                "communications", "wireless", "mobile", "cellular",
                "network", "broadband", "data", "fiber", "fibre"
            ],
            "bad": [
                "bank", "insurance", "mining", "property", "reit",
                "hotel", "lodging", "restaurants", "packaged foods"
            ],
        }

    if tgt_sector == "banking":
        return {
            "good": [
                "bank", "banking", "commercial bank", "retail bank",
                "lending", "deposits", "financial services"
            ],
            "bad": [
                "insurance", "mining", "telecom", "property", "reit",
                "hotel", "lodging", "restaurants"
            ],
        }

    if tgt_sector == "mining" and tgt_industry == "gold mining":
        return {
            "good": [
                "gold", "gold mining", "gold producer", "gold miner",
                "precious metals", "mining", "minerals", "resources"
            ],
            "bad": [
                "bank", "insurance", "telecom", "property", "reit",
                "hotel", "lodging", "restaurants", "packaged foods"
            ],
        }

    if tgt_sector == "mining":
        return {
            "good": [
                "mining", "minerals", "resources", "metals", "gold",
                "platinum", "copper", "ore", "exploration",
                "precious metals", "diversified mining"
            ],
            "bad": [
                "bank", "insurance", "telecom", "property", "reit",
                "hotel", "lodging", "restaurants", "packaged foods"
            ],
        }

    return {"good": [], "bad": []}
def strict_universe_filter(target_profile: dict, max_peers: int = 8):
    if UNIVERSE_DF is None or UNIVERSE_DF.empty:
        return pd.DataFrame()

    df = UNIVERSE_DF.copy()

    tgt_symbol = _clean_text(target_profile.get("target_symbol")).upper()
    tgt_company = _clean_text(target_profile.get("target_company"))
    tgt_sector = _clean_text(target_profile.get("preferred_sector")).lower()
    tgt_industry = _clean_text(target_profile.get("preferred_industry")).lower()
    tgt_keywords = [k.lower() for k in target_profile.get("preferred_peer_keywords", [])]

    df = df[df["ticker"].map(lambda x: _clean_text(x).upper()) != tgt_symbol].copy()
    df = df[df["company"].map(lambda x: _norm_text(x)) != _norm_text(tgt_company)].copy()

    if "yahoo_status" in df.columns:
        def status_ok(x):
            s = _clean_text(x).lower()
            if s == "":
                return True
            good_tokens = [
                "confirm", "likely", "yes", "active", "ok",
                "candidate", "valid", "listed", "yahoo"
            ]
            return any(tok in s for tok in good_tokens)

        df_status = df[df["yahoo_status"].map(status_ok)].copy()
        if not df_status.empty:
            df = df_status

    df["sector_l"] = df["sector"].map(lambda x: _clean_text(x).lower())
    df["industry_l"] = df["industry"].map(lambda x: _clean_text(x).lower())
    df["keywords_l"] = df["sector_keywords"].map(lambda x: _clean_text(x).lower())
    df["company_l"] = df["company"].map(lambda x: _clean_text(x).lower())

    fam = _family_words(target_profile)
    good_words = fam["good"]
    bad_words = fam["bad"]

    hit_cols = [
        "combo_text",
        "exact_sector_hit",
        "exact_industry_hit",
        "keyword_hits",
        "good_hits",
        "bad_hits",
    ]

    def compute_hits(row):
        combo = f"{row['sector_l']} | {row['industry_l']} | {row['keywords_l']} | {row['company_l']}"

        exact_sector = int(bool(
            tgt_sector and (
                row["sector_l"] == tgt_sector
                or tgt_sector in row["sector_l"]
                or row["sector_l"] in tgt_sector
            )
        ))

        exact_industry = int(bool(
            tgt_industry and (
                row["industry_l"] == tgt_industry
                or tgt_industry in row["industry_l"]
                or row["industry_l"] in tgt_industry
            )
        ))

        keyword_hits = sum(1 for kw in tgt_keywords if kw and kw in combo)
        good_hits = sum(1 for w in good_words if w in combo)
        bad_hits = sum(1 for w in bad_words if w in combo)

        return pd.Series({
            "combo_text": combo,
            "exact_sector_hit": exact_sector,
            "exact_industry_hit": exact_industry,
            "keyword_hits": keyword_hits,
            "good_hits": good_hits,
            "bad_hits": bad_hits,
        })

    if df.empty:
        for c in hit_cols:
            df[c] = pd.Series(dtype="object" if c == "combo_text" else "int")
    else:
        hit_df = df.apply(compute_hits, axis=1)

        if isinstance(hit_df, pd.Series):
            hit_df = pd.DataFrame(list(hit_df), index=df.index)

        for c in hit_cols:
            if c not in hit_df.columns:
                hit_df[c] = "" if c == "combo_text" else 0

        hit_df = hit_df[hit_cols]

        for c in hit_cols:
            df[c] = hit_df[c].values

    for c in ["exact_sector_hit", "exact_industry_hit", "keyword_hits", "good_hits", "bad_hits"]:
        if c not in df.columns:
            df[c] = 0

    strict_df = df[
        (
            (df["exact_sector_hit"] > 0) |
            (df["exact_industry_hit"] > 0) |
            (df["keyword_hits"] >= 2) |
            (df["good_hits"] >= 2)
        )
        &
        (
            (df["bad_hits"] == 0) | (df["good_hits"] >= 3)
        )
    ].copy()

    relaxed_df = df[
        (
            (df["exact_sector_hit"] > 0) |
            (df["exact_industry_hit"] > 0) |
            (df["keyword_hits"] >= 1) |
            (df["good_hits"] >= 1)
        )
        &
        (
            (df["bad_hits"] == 0) | (df["good_hits"] >= 2)
        )
    ].copy()

    chosen = strict_df if not strict_df.empty else relaxed_df

    if chosen.empty:
        S["debug_filter_stage_counts"] = {
            "universe_rows": len(df),
            "strict_rows": len(strict_df),
            "relaxed_rows": len(relaxed_df),
        }
        S["debug_strict_candidates_preview"] = df[[
            "ticker", "company", "sector", "industry", "sector_keywords",
            "exact_sector_hit", "exact_industry_hit", "keyword_hits", "good_hits", "bad_hits"
        ]].head(30)
        return chosen

    chosen["SimilarityScore"] = chosen.apply(lambda r: strict_peer_score(r.to_dict(), target_profile), axis=1)
    chosen = chosen[chosen["SimilarityScore"] > 0].copy()

    if chosen.empty:
        return chosen

    if "match_priority" not in chosen.columns:
        chosen["match_priority"] = ""

    chosen["match_priority_num"] = pd.to_numeric(chosen["match_priority"], errors="coerce").fillna(0)

    chosen = chosen.sort_values(
        by=["SimilarityScore", "match_priority_num", "company"],
        ascending=[False, False, True]
    ).drop_duplicates(subset=["ticker"]).reset_index(drop=True)

    S["debug_filter_stage_counts"] = {
        "universe_rows": len(df),
        "strict_rows": len(strict_df),
        "relaxed_rows": len(relaxed_df),
        "chosen_rows": len(chosen),
    }
    S["debug_strict_candidates_preview"] = chosen[[
        "ticker", "company", "sector", "industry", "sector_keywords",
        "exact_sector_hit", "exact_industry_hit", "keyword_hits", "good_hits", "bad_hits", "SimilarityScore"
    ]].head(30)

    return chosen.head(max(max_peers * 4, 16)).reset_index(drop=True)
# =========================================================
# LIVE RATIOS
# =========================================================
SESSION = requests.Session()

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json,text/plain,*/*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://finance.yahoo.com/",
    "Origin": "https://finance.yahoo.com",
    "X-Requested-With": "XMLHttpRequest",
    "Connection": "keep-alive",
}

YAHOO_QUOTESUMMARY_URL = "https://query2.finance.yahoo.com/v10/finance/quoteSummary/{symbol}"
YAHOO_QUOTE_URL = "https://query1.finance.yahoo.com/v7/finance/quote"


def _safe_get(url, params=None, timeout=25, tries=3, headers=None):
    last_err = None
    use_headers = headers or HEADERS

    for i in range(int(tries)):
        try:
            r = SESSION.get(url, params=params, timeout=timeout, headers=use_headers)
            if r.status_code == 429:
                continue
            r.raise_for_status()
            return r
        except Exception as e:
            last_err = e
    raise last_err


def yahoo_warmup():
    try:
        _safe_get("https://finance.yahoo.com/", timeout=15, tries=2, headers={
            "User-Agent": HEADERS["User-Agent"],
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": "https://finance.yahoo.com/",
        })
        _safe_get(
            "https://query2.finance.yahoo.com/v1/finance/search",
            params={"q": "test", "quotesCount": 1, "newsCount": 0},
            timeout=15,
            tries=2,
        )
    except Exception:
        pass


yahoo_warmup()


def make_yahoo_statistics_url(symbol: str) -> str:
    sym = normalize_peer_ticker(symbol)
    return f"https://finance.yahoo.com/quote/{sym}/key-statistics?p={sym}" if sym else ""


def _first_non_null(d: dict, keys):
    for k in keys:
        v = d.get(k)
        if v is not None and v != "":
            return v
    return None


@st.cache_data(show_spinner=False, ttl=60 * 60 * 6)
def yahoo_profile_and_metrics(symbol: str) -> dict:
    sym = normalize_peer_ticker(symbol)
    if not sym:
        return {}

    def _raw(v):
        if isinstance(v, dict):
            if "raw" in v:
                return v.get("raw")
            if "fmt" in v:
                return v.get("fmt")
        return v

    out = {
        "Ticker": sym,
        "Company": "",
        "Exchange": "",
        "Country": "",
        "Sector": "",
        "Industry": "",
        "Description": "",
        "P/E": np.nan,
        "P/B": np.nan,
        "EV/EBITDA": np.nan,
        "ratio_source": "",
    }

    try:
        url = YAHOO_QUOTESUMMARY_URL.format(symbol=sym)
        r = _safe_get(
            url,
            params={
                "modules": "price,summaryDetail,defaultKeyStatistics,financialData,assetProfile"
            },
            timeout=25,
            tries=2,
        )
        data = r.json()
        res = (((data.get("quoteSummary") or {}).get("result")) or [])
        if not res:
            return out

        root = res[0]
        price = root.get("price") or {}
        summary = root.get("summaryDetail") or {}
        dks = root.get("defaultKeyStatistics") or {}
        fin = root.get("financialData") or {}
        ap = root.get("assetProfile") or {}

        out["Company"] = (
            _raw(price.get("longName"))
            or _raw(price.get("shortName"))
            or sym
        )
        out["Exchange"] = _raw(price.get("exchangeName")) or _raw(price.get("fullExchangeName")) or ""
        out["Country"] = ap.get("country") or ""
        out["Sector"] = ap.get("sector") or ""
        out["Industry"] = ap.get("industry") or ""
        out["Description"] = ap.get("longBusinessSummary") or ""

        pe = _raw(summary.get("trailingPE")) or _raw(dks.get("trailingPE")) or _raw(fin.get("trailingPE"))
        pb = _raw(dks.get("priceToBook")) or _raw(summary.get("priceToBook")) or _raw(fin.get("priceToBook"))
        evebitda = _raw(fin.get("enterpriseToEbitda")) or _raw(dks.get("enterpriseToEbitda"))

        out["P/E"] = _clean_num(pe)
        out["P/B"] = _clean_num(pb)
        out["EV/EBITDA"] = _clean_num(evebitda)

        if not (pd.isna(out["P/E"]) and pd.isna(out["P/B"]) and pd.isna(out["EV/EBITDA"])):
            out["ratio_source"] = "Yahoo quoteSummary"
    except Exception:
        pass

    return out


@st.cache_data(show_spinner=False, ttl=60 * 60 * 6)
def yahoo_stats_table_fallback(symbol: str) -> dict:
    sym = normalize_peer_ticker(symbol)

    out = {
        "Ticker": sym,
        "P/E": np.nan,
        "P/B": np.nan,
        "EV/EBITDA": np.nan,
        "ratio_source": "",
        "ratio_note": "",
    }

    if not sym:
        out["ratio_note"] = "Blank symbol after normalization."
        return out

    def parse_ratio_value(x):
        s = str(x).strip()
        if s in ["", "N/A", "NaN", "None", "-", "--"]:
            return np.nan

        s = s.replace(",", "").strip()
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        if not m:
            return np.nan

        try:
            return float(m.group(0))
        except Exception:
            return np.nan

    url = f"https://finance.yahoo.com/quote/{sym}/key-statistics?p={sym}"

    html_headers = {
        "User-Agent": HEADERS["User-Agent"],
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": "https://finance.yahoo.com/",
    }

    try:
        r = _safe_get(url, timeout=25, tries=2, headers=html_headers)
        html = r.text or ""

        soup = BeautifulSoup(html, "html.parser")

        # 1) Parse actual HTML tables exactly like the Statistics page
        for table in soup.find_all("table"):

            rows = table.find_all("tr")
            if not rows:
                continue

            # detect header row
            header_cells = rows[0].find_all(["th", "td"])
            headers = [c.get_text(" ", strip=True).lower() for c in header_cells]

            current_idx = None
            for i, h in enumerate(headers):
                if "current" in h:
                    current_idx = i
                    break

            forward_pe_val = np.nan
            trailing_pe_val = np.nan
            pb_val = np.nan
            ev_ebitda_val = np.nan

            for tr in rows[1:]:
                cells = tr.find_all(["th", "td"])
                vals = [c.get_text(" ", strip=True) for c in cells]

                if len(vals) < 2:
                    continue

                label = vals[0].strip().lower()

                if current_idx is not None and len(vals) > current_idx:
                    value = vals[current_idx]
                else:
                    value = vals[-1]

                if "forward p/e" in label and pd.isna(forward_pe_val):
                    forward_pe_val = parse_ratio_value(value)

                elif "trailing p/e" in label and pd.isna(trailing_pe_val):
                    trailing_pe_val = parse_ratio_value(value)

                elif ("price/book" in label or "price to book" in label) and pd.isna(pb_val):
                    pb_val = parse_ratio_value(value)

                elif ("enterprise value/ebitda" in label or "ev/ebitda" in label) and pd.isna(ev_ebitda_val):
                    ev_ebitda_val = parse_ratio_value(value)

            # apply priority after scanning full table
            if pd.isna(out["P/E"]):
                if not pd.isna(forward_pe_val):
                    out["P/E"] = forward_pe_val
                elif not pd.isna(trailing_pe_val):
                    out["P/E"] = trailing_pe_val

            if pd.isna(out["P/B"]) and not pd.isna(pb_val):
                out["P/B"] = pb_val

            if pd.isna(out["EV/EBITDA"]) and not pd.isna(ev_ebitda_val):
                out["EV/EBITDA"] = ev_ebitda_val

                if pd.isna(out["P/E"]) and "trailing p/e" in label:
                    out["P/E"] = parse_ratio_value(value)

                elif pd.isna(out["P/B"]) and ("price/book" in label or "price to book" in label):
                    out["P/B"] = parse_ratio_value(value)

                elif pd.isna(out["EV/EBITDA"]) and (
                    "enterprise value/ebitda" in label or "ev/ebitda" in label
                ):
                    out["EV/EBITDA"] = parse_ratio_value(value)

        # 2) Fallback: pandas read_html on same page
        if pd.isna(out["P/E"]) or pd.isna(out["P/B"]) or pd.isna(out["EV/EBITDA"]):
            try:
                tables = pd.read_html(StringIO(html))
            except Exception:
                tables = []

            for t in tables:
                if t is None or t.empty:
                    continue

                for _, row in t.iterrows():
                    vals = [
                        str(x).strip()
                        for x in row.tolist()
                        if str(x).strip() not in ["", "nan", "None"]
                    ]
                    if len(vals) < 2:
                        continue

                    label = vals[0].lower()
                    value = vals[-1]

                    if pd.isna(out["P/E"]) and "trailing p/e" in label:
                        out["P/E"] = parse_ratio_value(value)

                    elif pd.isna(out["P/B"]) and ("price/book" in label or "price to book" in label):
                        out["P/B"] = parse_ratio_value(value)

                    elif pd.isna(out["EV/EBITDA"]) and (
                        "enterprise value/ebitda" in label or "ev/ebitda" in label
                    ):
                        out["EV/EBITDA"] = parse_ratio_value(value)

        # 3) Last fallback: regex directly from page source
        if pd.isna(out["P/E"]):
            m = re.search(r"Trailing P/E.*?(-?\d+(?:\.\d+)?)", html, re.I | re.S)
            if m:
                out["P/E"] = parse_ratio_value(m.group(1))

        if pd.isna(out["P/B"]):
            m = re.search(r"Price/Book.*?(-?\d+(?:\.\d+)?)", html, re.I | re.S)
            if m:
                out["P/B"] = parse_ratio_value(m.group(1))

        if pd.isna(out["EV/EBITDA"]):
            m = re.search(r"Enterprise Value/EBITDA.*?(-?\d+(?:\.\d+)?)", html, re.I | re.S)
            if m:
                out["EV/EBITDA"] = parse_ratio_value(m.group(1))

        if not (pd.isna(out["P/E"]) and pd.isna(out["P/B"]) and pd.isna(out["EV/EBITDA"])):
            out["ratio_source"] = "Yahoo Statistics"
            out["ratio_note"] = "Fetched from Yahoo Statistics"
        else:
            out["ratio_note"] = "Yahoo stats page loaded but ratio rows were not found."

    except Exception as e:
        out["ratio_note"] = f"Yahoo stats fetch failed: {repr(e)}"

    return out


@st.cache_data(show_spinner=False, ttl=60 * 60 * 6)
def yahoo_html_ratio_fallback(symbol: str) -> dict:
    sym = normalize_peer_ticker(symbol)
    out = {
        "Ticker": sym,
        "P/E": np.nan,
        "P/B": np.nan,
        "EV/EBITDA": np.nan,
        "ratio_source": ""
    }

    if not sym:
        return out

    try:
        url = f"https://finance.yahoo.com/quote/{sym}"
        r = _safe_get(url, timeout=20, tries=2)
        html = r.text or ""

        patterns = {
            "P/E": [
                r'"trailingPE"\s*:\s*\{"raw"\s*:\s*([\-0-9.]+)',
                r'"trailingPE"\s*:\s*([\-0-9.]+)',
            ],
            "P/B": [
                r'"priceToBook"\s*:\s*\{"raw"\s*:\s*([\-0-9.]+)',
                r'"priceToBook"\s*:\s*([\-0-9.]+)',
            ],
            "EV/EBITDA": [
                r'"enterpriseToEbitda"\s*:\s*\{"raw"\s*:\s*([\-0-9.]+)',
                r'"enterpriseToEbitda"\s*:\s*([\-0-9.]+)',
            ],
        }

        for field, pats in patterns.items():
            for pat in pats:
                m = re.search(pat, html)
                if m:
                    out[field] = _clean_num(m.group(1))
                    break

        if not (pd.isna(out["P/E"]) and pd.isna(out["P/B"]) and pd.isna(out["EV/EBITDA"])):
            out["ratio_source"] = "Yahoo HTML fallback"
    except Exception:
        pass

    return out


@st.cache_data(show_spinner=False, ttl=60 * 60 * 6)
def get_live_peer_row(
    symbol: str,
    fallback_company: str = "",
    fallback_country: str = "",
    fallback_exchange: str = "",
    fallback_sector: str = "",
    fallback_industry: str = ""
):
    sym = normalize_peer_ticker(symbol)

    out = {
        "Company": fallback_company or sym,
        "Ticker": sym,
        "Exchange": fallback_exchange,
        "Country": fallback_country,
        "Sector": fallback_sector,
        "Industry": fallback_industry,
        "EV/EBITDA": np.nan,
        "P/B": np.nan,
        "P/E": np.nan,
        "Source": "",
        "RatioNote": "",
        "YahooProfile": make_yahoo_profile_url(sym),
        "YahooStats": make_yahoo_statistics_url(sym),
    }

    if not sym:
        return out

    # 1) Yahoo quoteSummary + profile
    yh = yahoo_profile_and_metrics(sym)

    # 2) Yahoo Statistics tab fallback
    ystats = yahoo_stats_table_fallback(sym)

    # 3) Yahoo HTML fallback
    yhtml = yahoo_html_ratio_fallback(sym)

    # 4) yfinance only for company/profile details
    info = {}
    try:
        tk = yf.Ticker(sym)
        info = tk.info or {}
    except Exception:
        info = {}

    company = (
        _clean_text(yh.get("Company"))
        or _clean_text(info.get("longName") or info.get("shortName"))
        or fallback_company
        or sym
    )
    exchange = (
        _clean_text(yh.get("Exchange"))
        or _clean_text(info.get("exchange"))
        or fallback_exchange
    )
    country = (
        _clean_text(yh.get("Country"))
        or _clean_text(info.get("country"))
        or fallback_country
    )
    sector = (
        _clean_text(yh.get("Sector"))
        or _clean_text(info.get("sector"))
        or fallback_sector
    )
    industry = (
        _clean_text(yh.get("Industry"))
        or _clean_text(info.get("industry"))
        or fallback_industry
    )

    # Ratio priority = same working logic
    pe = ystats.get("P/E", np.nan)
    if pd.isna(pe):
        pe = yh.get("P/E", np.nan)
    if pd.isna(pe):
        pe = yhtml.get("P/E", np.nan)

    pb = ystats.get("P/B", np.nan)
    if pd.isna(pb):
        pb = yh.get("P/B", np.nan)
    if pd.isna(pb):
        pb = yhtml.get("P/B", np.nan)

    ev_ebitda = ystats.get("EV/EBITDA", np.nan)
    if pd.isna(ev_ebitda):
        ev_ebitda = yh.get("EV/EBITDA", np.nan)
    if pd.isna(ev_ebitda):
        ev_ebitda = yhtml.get("EV/EBITDA", np.nan)

    source = (
        ystats.get("ratio_source")
        or yh.get("ratio_source")
        or yhtml.get("ratio_source")
        or ""
    )

    ratio_note = (
        ystats.get("ratio_note")
        or ("Fetched from Yahoo quoteSummary" if yh.get("ratio_source") else "")
        or ("Fetched from Yahoo HTML fallback" if yhtml.get("ratio_source") else "")
        or "No Yahoo ratio source returned values."
    )

    out.update({
        "Company": company,
        "Exchange": exchange,
        "Country": country,
        "Sector": sector,
        "Industry": industry,
        "EV/EBITDA": _clean_num(ev_ebitda),
        "P/B": _clean_num(pb),
        "P/E": _clean_num(pe),
        "Source": source,
        "RatioNote": ratio_note,
    })

    return out
def build_live_comps_from_target(target_query: str, max_peers: int = 8, manual_sector_override: str = ""):
    target_profile = get_target_profile(target_query, manual_sector_override)

    strict_df = strict_universe_filter(target_profile, max_peers=max_peers)
    S["debug_strict_df_shape"] = strict_df.shape if strict_df is not None else (0, 0)
    S["debug_strict_df_preview"] = strict_df.head(20) if strict_df is not None and not strict_df.empty else pd.DataFrame()

    if strict_df is None or strict_df.empty:
        return pd.DataFrame(), {
            "error": f"No peers found in the strict Africa universe for {target_profile.get('target_symbol') or target_query}.",
            "target": target_profile,
            "peer_source": "Africa universe Excel",
        }

    rows = []
    for _, r in strict_df.iterrows():
        live = get_live_peer_row(
            symbol=r.get("ticker", ""),
            fallback_company=r.get("company", ""),
            fallback_country=r.get("country", ""),
            fallback_exchange=r.get("exchange", ""),
            fallback_sector=r.get("sector", ""),
            fallback_industry=r.get("industry", ""),
        )
        live["SimilarityScore"] = strict_peer_score(r.to_dict(), target_profile)
        live["UniverseSector"] = _clean_text(r.get("sector"))
        live["UniverseIndustry"] = _clean_text(r.get("industry"))
        live["UniverseKeywords"] = _clean_text(r.get("sector_keywords"))
        rows.append(live)

    df = pd.DataFrame(rows).drop_duplicates(subset=["Ticker"]).reset_index(drop=True)

    if df.empty:
        return pd.DataFrame(), {
            "error": f"Peers matched in Excel, but live ratio fetch returned no rows for {target_profile.get('target_symbol') or target_query}.",
            "target": target_profile,
            "peer_source": "Africa universe Excel",
        }

    df["RatioCount"] = (
        df["EV/EBITDA"].notna().astype(int)
        + df["P/B"].notna().astype(int)
        + df["P/E"].notna().astype(int)
    )

    df = df.sort_values(
        by=["SimilarityScore", "RatioCount", "Company"],
        ascending=[False, False, True]
    ).head(max_peers).reset_index(drop=True)

    meta = {
        "target": target_profile,
        "peer_source": "Africa universe Excel",
        "peer_count": len(df),
        "target_sector": target_profile.get("preferred_sector", ""),
        "target_industry": target_profile.get("preferred_industry", ""),
    }

    return df, meta
def apply_live_comps_to_session(df_live: pd.DataFrame):
    if df_live is None or df_live.empty:
        return

    S.setdefault("comps", {})

    n = len(df_live)
    S["num_comps"] = n

    for i, (_, r) in enumerate(df_live.iterrows()):
        S[f"comp_name_{i}"] = _clean_text(r.get("Company")) or _clean_text(r.get("Ticker"))
        S[f"comp_ticker_{i}"] = _clean_text(r.get("Ticker"))
        S[f"comp_source_{i}"] = _clean_text(r.get("Source"))
        S[f"comp_profile_{i}"] = _clean_text(r.get("YahooProfile"))

        S[f"comp_ev_{i}"] = np.nan if pd.isna(r.get("EV/EBITDA")) else float(r["EV/EBITDA"])
        S[f"comp_pb_{i}"] = np.nan if pd.isna(r.get("P/B")) else float(r["P/B"])
        S[f"comp_pe_{i}"] = np.nan if pd.isna(r.get("P/E")) else float(r["P/E"])

        S[f"inc_ev_{i}"] = not pd.isna(r.get("EV/EBITDA"))
        S[f"inc_pb_{i}"] = not pd.isna(r.get("P/B"))
        S[f"inc_pe_{i}"] = not pd.isna(r.get("P/E"))

        S["comps"].setdefault(i, {})
        S["comps"][i]["name"] = S[f"comp_name_{i}"]
        S["comps"][i]["ticker"] = S[f"comp_ticker_{i}"]
        S["comps"][i]["source"] = S[f"comp_source_{i}"]
        S["comps"][i]["profile"] = S[f"comp_profile_{i}"]
        S["comps"][i]["ev"] = S[f"comp_ev_{i}"]
        S["comps"][i]["pb"] = S[f"comp_pb_{i}"]
        S["comps"][i]["pe"] = S[f"comp_pe_{i}"]
        S["comps"][i]["inc_ev"] = S[f"inc_ev_{i}"]
        S["comps"][i]["inc_pb"] = S[f"inc_pb_{i}"]
        S["comps"][i]["inc_pe"] = S[f"inc_pe_{i}"]


# =========================================================
# STEP 1 — INPUT COMPARABLE COMPANIES & MULTIPLES
# =========================================================
st.header("Step 1 — Input Comparable Companies & Multiples")
st.subheader("Auto Peer Suggestions from Strict Africa Universe Excel")

if UNIVERSE_FILE:
    st.caption(f"Universe file loaded: {UNIVERSE_FILE}")
else:
    st.error("❌ Africa peer universe Excel was not found. Put it inside a data folder, e.g. data/africa_yahoo_peer_universe_strict_final.xlsx")
    st.stop()

S.setdefault("target_company", "")
S.setdefault("auto_peer_count", 8)
S.setdefault("manual_sector_override", "")
S.setdefault("live_comps_df", pd.DataFrame())
S.setdefault("live_comps_meta", {})

cA, cB, cC = st.columns([2.2, 1, 1.2])
with cA:
    target_company = st.text_input(
        "Company you are valuing (Zimbabwe / VFEX / JSE / any ticker in your universe)",
        value=S["target_company"],
        key="target_company_input",
        placeholder="e.g. econet, padenga, cbz, delta, innscor ...",
    )
with cB:
    peer_count = st.number_input(
        "Peers to suggest",
        min_value=3,
        max_value=15,
        value=int(S["auto_peer_count"]),
        step=1,
        key="auto_peer_count_input",
    )
with cC:
    st.caption(" ")
    auto_apply = st.checkbox("Auto-fill Step 1 names", value=True, key="auto_apply_peers")

S["target_company"] = target_company
S["auto_peer_count"] = int(peer_count)

manual_sector = st.text_input(
    "Optional manual sector override",
    value=S["manual_sector_override"],
    key="manual_sector_override_input",
    placeholder="e.g. telecommunications, banking, mining, beverages",
)
S["manual_sector_override"] = manual_sector

st.subheader("Live peer search and ratio fill")
st.caption("Peers come from your Africa universe Excel first, then ratios are fetched from Yahoo Finance Statistics tab.")

live_peer_limit = st.slider(
    "Live peers to import",
    min_value=3,
    max_value=12,
    value=min(int(S["auto_peer_count"]), 12),
    step=1,
    key="live_peer_limit"
)

run_live_comps = st.button("⚡ Auto-search live peers and ratios")
if st.button("Clear ratio cache"):
    st.cache_data.clear()
    st.success("Cache cleared. Run live peer search again.")
if run_live_comps:
    if not target_company.strip():
        st.warning("Enter the company name or ticker first.")
    else:
        with st.spinner("Searching strict Excel peers and live ratios..."):
            live_df, meta = build_live_comps_from_target(
                target_query=target_company,
                max_peers=int(live_peer_limit),
                manual_sector_override=manual_sector,
            )
            S["live_comps_df"] = live_df
            S["live_comps_meta"] = meta

            if live_df is not None and not live_df.empty:
                if auto_apply:
                    apply_live_comps_to_session(live_df)
                st.success(f"Loaded {len(live_df)} live peers.")
            else:
                st.error(meta.get("error", "Live peer search failed."))

live_df = S.get("live_comps_df", pd.DataFrame())
live_meta = S.get("live_comps_meta", {})

if live_meta:
    tgt = live_meta.get("target", {})
    st.caption(
        f"Resolved target: {tgt.get('target_company', '')} ({tgt.get('target_symbol', '')}) "
        f"via {tgt.get('source', '')} | Peer source: {live_meta.get('peer_source', '')} "
        f"| Manual sector override: {S.get('manual_sector_override', '') or 'None'}"
    )

if live_df is not None and not live_df.empty:
    df_show = live_df.copy()
    ratio_cols = ["EV/EBITDA", "P/B", "P/E"]

    for c in ratio_cols:
        df_show[c] = pd.to_numeric(df_show[c], errors="coerce")

    display_cols = [
        "YahooStats",
        "YahooProfile",
        "Company",
        "Ticker",
        "Exchange",
        "Country",
        "Sector",
        "Industry",
        "SimilarityScore",
        "EV/EBITDA",
        "P/B",
        "P/E",
        "Source",
    ]
    st.dataframe(
        df_show[display_cols],
        use_container_width=True,
        column_config={
            "YahooStats": st.column_config.LinkColumn("Yahoo Statistics", display_text="Open Stats"),
            "YahooProfile": st.column_config.LinkColumn("Yahoo Profile", display_text="Open"),
            "SimilarityScore": st.column_config.NumberColumn("SimilarityScore", format="%d"),
            "EV/EBITDA": st.column_config.NumberColumn("EV/EBITDA", format="%.2f"),
            "P/B": st.column_config.NumberColumn("P/B", format="%.2f"),
            "P/E": st.column_config.NumberColumn("P/E", format="%.2f"),
        }
    )

    missing_all = df_show[ratio_cols].isna().all(axis=1)
    if missing_all.any():
        st.warning(
            "Some peers were found, but some ratio fields are still unavailable from Yahoo Finance Statistics tab.")

with st.expander("Debug peer search"):
    st.write("Universe debug:", UNIVERSE_DEBUG)
    if live_meta:
        st.write("Target profile:", live_meta.get("target", {}))
    st.write("Strict df shape:", S.get("debug_strict_df_shape"))
    st.write("Filter stage counts:", S.get("debug_filter_stage_counts", {}))
    st.write("Strict df preview:", S.get("debug_strict_df_preview"))
    st.write("Candidate preview:", S.get("debug_strict_candidates_preview", pd.DataFrame()))
    if live_df is not None and not live_df.empty:
        st.write("Returned peer tickers:", live_df["Ticker"].tolist())

        debug_cols = ["Ticker", "Source", "RatioNote", "EV/EBITDA", "P/B", "P/E"]
        available_debug_cols = [c for c in debug_cols if c in live_df.columns]
        st.dataframe(live_df[available_debug_cols], use_container_width=True)
        st.write("Ratio notes by ticker:")
        for _, rr in live_df.iterrows():
            st.write(f"{rr.get('Ticker', '')}: {rr.get('RatioNote', '')}")
# =========================================================
# STEP 1 MANUAL COMPS INPUT
# =========================================================
S.setdefault("num_comps", 3)
S.setdefault("comps", {})

num_comps = st.number_input(
    "How many comparables?",
    min_value=1,
    max_value=20,
    value=int(S.get("num_comps", 3)),
    key="num_comps_input",
)
S["num_comps"] = int(num_comps)

for i in range(int(num_comps)):
    S["comps"].setdefault(i, {
        "name": f"Comp {i + 1}",
        "ev": np.nan,
        "pb": np.nan,
        "pe": np.nan,
        "inc_ev": True,
        "inc_pb": True,
        "inc_pe": True,
    })

rows = []
for i in range(int(num_comps)):
    st.subheader(f"Comparable {i + 1}")
    c1, c2, c3, c4, c5 = st.columns([2, 1, 1, 1, 1.2])

    with c1:
        default_name = S.get(f"comp_name_{i}", S["comps"][i]["name"])
        name = st.text_input(
            f"Company {i + 1} name",
            value=str(default_name),
            key=f"comp_name_{i}",
        )
        S["comps"][i]["name"] = name

    with c2:
        default_ev = _num_input_default(S.get(f"comp_ev_{i}", S["comps"][i]["ev"]), 0.0)
        ev = st.number_input(
            f"{name} EV/EBITDA",
            value=default_ev,
            step=0.01,
            format="%.2f",
            key=f"comp_ev_{i}",
        )
        S["comps"][i]["ev"] = ev

    with c3:
        default_pb = _num_input_default(S.get(f"comp_pb_{i}", S["comps"][i]["pb"]), 0.0)
        pb = st.number_input(
            f"{name} P/B",
            value=default_pb,
            step=0.01,
            format="%.2f",
            key=f"comp_pb_{i}",
        )
        S["comps"][i]["pb"] = pb

    with c4:
        default_pe = _num_input_default(S.get(f"comp_pe_{i}", S["comps"][i]["pe"]), 0.0)
        pe = st.number_input(
            f"{name} P/E",
            value=default_pe,
            step=0.01,
            format="%.2f",
            key=f"comp_pe_{i}",
        )
        S["comps"][i]["pe"] = pe

    ev_key = f"inc_ev_{i}"
    pb_key = f"inc_pb_{i}"
    pe_key = f"inc_pe_{i}"

    if ev_key not in S:
        S[ev_key] = bool(S["comps"][i].get("inc_ev", True))
    if pb_key not in S:
        S[pb_key] = bool(S["comps"][i].get("inc_pb", True))
    if pe_key not in S:
        S[pe_key] = bool(S["comps"][i].get("inc_pe", True))

    with c5:
        st.caption("Analyst filter")
        st.checkbox("Include EV/EBITDA", key=ev_key)
        st.checkbox("Include P/B", key=pb_key)
        st.checkbox("Include P/E", key=pe_key)

    ticker_val = S.get(f"comp_ticker_{i}", "")
    source_val = S.get(f"comp_source_{i}", "")
    profile_val = S.get(f"comp_profile_{i}", "")

    if ticker_val or source_val:
        st.caption(f"Ticker: {ticker_val} | Ratio source: {source_val}")
        if profile_val:
            st.markdown(f"[Open Yahoo profile for {name}]({profile_val})")

    inc_ev = bool(S[ev_key])
    inc_pb = bool(S[pb_key])
    inc_pe = bool(S[pe_key])

    S["comps"][i]["inc_ev"] = inc_ev
    S["comps"][i]["inc_pb"] = inc_pb
    S["comps"][i]["inc_pe"] = inc_pe

    rows.append([name, ev, pb, pe, inc_ev, inc_pb, inc_pe])

df_comps = pd.DataFrame(
    rows,
    columns=["Company", "EV/EBITDA", "P/B", "P/E", "Include_EV", "Include_PB", "Include_PE"]
)

st.subheader("Entered Comparables")
st.dataframe(df_comps, use_container_width=True)

S["comps_num"] = int(num_comps)
S["comps_ev_list"] = df_comps["EV/EBITDA"].astype(float).tolist()
S["comps_pb_list"] = df_comps["P/B"].astype(float).tolist()
S["comps_pe_list"] = df_comps["P/E"].astype(float).tolist()
S["comps_inc_ev"] = df_comps["Include_EV"].astype(bool).tolist()
S["comps_inc_pb"] = df_comps["Include_PB"].astype(bool).tolist()
S["comps_inc_pe"] = df_comps["Include_PE"].astype(bool).tolist()


# =========================================================
# STEP 2: AVERAGES
# =========================================================
st.header("Step 2 — Average & Implied Multiples")

ev_series = df_comps.loc[df_comps["Include_EV"] == True, "EV/EBITDA"]
pb_series = df_comps.loc[df_comps["Include_PB"] == True, "P/B"]
pe_series = df_comps.loc[df_comps["Include_PE"] == True, "P/E"]

avg_ev = filtered_average(ev_series)
avg_pb = filtered_average(pb_series)
avg_pe = filtered_average(pe_series)

discount_pct = st.number_input(
    "Discount factor (%)",
    value=float(st.session_state.get("discount_pct", 25.0)),
    step=1.0,
    key="discount_pct",
)
discount = discount_pct / 100

implied_ev = avg_ev * (1 - discount)
implied_pb = avg_pb * (1 - discount)
implied_pe = avg_pe * (1 - discount)

st.dataframe(
    pd.DataFrame({
        "Multiple": ["EV/EBITDA", "P/B", "P/E"],
        "Average": [avg_ev, avg_pb, avg_pe],
        "Discount (%)": [discount_pct] * 3,
        "Implied": [implied_ev, implied_pb, implied_pe]
    }).style.format({"Average": "{:,.2f}", "Implied": "{:,.2f}"}),
    use_container_width=True
)

S["implied_ev"] = float(implied_ev) if not pd.isna(implied_ev) else 0.0
S["implied_pb"] = float(implied_pb) if not pd.isna(implied_pb) else 0.0
S["implied_pe"] = float(implied_pe) if not pd.isna(implied_pe) else 0.0

# =========================================================
# TIMING SOURCE (from DCF) — BASE USED BY BOTH EBITDA & EARNINGS
# =========================================================
st.header("Timing Source (from DCF)")

dcf_timing_list = S.get("dcf_discount_periods_n", [])
default_base = float(S.get("comp_timing_base", 0.0))

if not dcf_timing_list:
    st.warning(
        "⚠ No timing values detected from DCF. "
        "Either run the DCF model first or set a manual timing base."
    )
    base_timing = st.number_input(
        "Enter starting timing value for comparables (year 1):",
        value=default_base,
        step=0.01,
        format="%.4f",
        key="comp_timing_base_manual_no_dcf",
    )
else:
    timing_df = pd.DataFrame(
        {"Forecast Year Index": list(range(len(dcf_timing_list))), "DCF Timing n": dcf_timing_list}
    )
    st.dataframe(timing_df, width='stretch')

    dcf_n0 = float(round(dcf_timing_list[0], 4))
    st.info(f"DCF First Timing Value (n₀) = **{dcf_n0} years**")

    timing_choice = st.radio(
        "Choose timing base for Comparables timing effect:",
        [f"Use DCF n₀ = {dcf_n0}", "Manually override starting timing value"],
        index=0 if default_base == 0.0 or np.isclose(default_base, dcf_n0) else 1,
        key="comp_timing_choice",
    )

    if timing_choice.startswith("Use DCF"):
        base_timing = dcf_n0
    else:
        base_timing = st.number_input(
            "Enter starting timing value for comparables (year 1):",
            value=default_base if default_base != 0.0 else dcf_n0,
            step=0.01,
            format="%.4f",
            key="comp_timing_base_manual",
        )

S["comp_timing_base"] = float(base_timing)
st.success(f"Timing base for comparables = **{base_timing:.4f}**")

# =========================================================
# STEP 3 — MAINTAINABLE EBITDA (with locked timing)
# =========================================================
st.header("Step 3 — Maintainable EBITDA")

dcf_eb_all = S.get("dcf_ebitda_all", None)
if dcf_eb_all is None:
    dcf_eb_all = S.get("dcf_ebitda_forecast", {})

# ✅ If no DCF EBITDA → SKIP Step 3 (no manual inputs)
if not dcf_eb_all:
    st.warning("⚠ No EBITDA found from DCF — skipping EV/EBITDA method.")
    S["maintainable_ebitda"] = np.nan

else:
    # ✅ SAFETY: only accept 4-digit year keys (e.g., 2024)
    eb_years_all = sorted(
        int(y) for y in dcf_eb_all.keys()
        if str(y).strip().isdigit() and len(str(y).strip()) == 4
    )

    if not eb_years_all:
        st.warning("⚠ DCF EBITDA found, but no valid 4-digit year keys — skipping EV/EBITDA method.")
        S["maintainable_ebitda"] = np.nan

    else:
        eb_min_year = min(eb_years_all)
        eb_max_year = max(eb_years_all)

        S.setdefault("comp_eb_start_year", eb_min_year)
        S.setdefault("comp_eb_end_year", eb_max_year)
        S.setdefault("comp_eb_weights", {})
        S.setdefault("comp_use_timing_eb", True)

        use_timing_eb = st.checkbox(
            "Apply timing effect from DCF to EBITDA?",
            value=bool(S.get("comp_use_timing_eb", True)),
            key="comp_use_timing_eb_checkbox",
        )
        S["comp_use_timing_eb"] = use_timing_eb
        # ---------------------------------------------------------
        # ✅ HARD SYNC: Earnings timing ALWAYS follows EBITDA timing when EBITDA changes
        # ---------------------------------------------------------
        prev_eb = S.get("_prev_comp_use_timing_eb", None)

        # if EBITDA timing changed this run, force earnings timing to match
        if prev_eb is None or bool(prev_eb) != bool(use_timing_eb):
            S["comp_use_timing_np"] = bool(use_timing_eb)
            S["comp_use_timing_np_checkbox"] = bool(use_timing_eb)  # this updates the UI checkbox

        S["_prev_comp_use_timing_eb"] = bool(use_timing_eb)

        # If user turned off EBITDA timing, also turn off Earnings timing immediately
        if not use_timing_eb:
            S["comp_use_timing_np"] = False
            S["comp_use_timing_np_checkbox"] = False

        c_eb1, c_eb2 = st.columns(2)
        with c_eb1:
            eb_start_year = st.number_input(
                "EBITDA Start Year",
                value=int(S["comp_eb_start_year"]),
                step=1,
                key="comp_eb_start_year_input",
            )
        with c_eb2:
            eb_end_year = st.number_input(
                "EBITDA End Year",
                value=int(S["comp_eb_end_year"]),
                step=1,
                key="comp_eb_end_year_input",
            )

        eb_start_year = int(max(eb_start_year, eb_min_year))
        eb_end_year = int(min(eb_end_year, eb_max_year))
        if eb_end_year < eb_start_year:
            st.error("❌ EBITDA End Year must be ≥ Start Year.")
            st.stop()

        S["comp_eb_start_year"] = eb_start_year
        S["comp_eb_end_year"] = eb_end_year

        selected_eb_years = list(range(eb_start_year, eb_end_year + 1))
        st.subheader("EBITDA Weighting")

        rows_eb = []
        base_timing = float(S.get("comp_timing_base", 0.0))

        for idx, yr in enumerate(selected_eb_years):
            eb_val = float(dcf_eb_all.get(str(yr), 0.0))
            default_w = float(S["comp_eb_weights"].get(str(yr), 0.0))

            if not use_timing_eb:
                timing_val = 1.0
            else:
                timing_val = base_timing + idx

            c1, c2, c4 = st.columns([1, 2, 1])
            with c1:
                st.number_input(f"EB Year {yr}", value=int(yr), disabled=True, key=f"comp_eb_year_display_{yr}")
            with c2:
                st.number_input(f"EBITDA {yr}", value=eb_val, disabled=True, format="%.2f", key=f"comp_eb_value_display_{yr}")
            with c4:
                weight_val = st.number_input(
                    f"EB Weight {yr} (%)",
                    value=float(default_w),
                    step=0.1,
                    format="%.2f",
                    key=f"comp_eb_weight_{yr}",
                )

            S["comp_eb_weights"][str(yr)] = float(weight_val)

            adj_eb = eb_val * timing_val
            weighted_eb = adj_eb * weight_val / 100.0

            rows_eb.append(
                {
                    "Year": int(yr),
                    "EBITDA": eb_val,
                    "Timing": timing_val if use_timing_eb else np.nan,
                    "Weight (%)": weight_val,
                    "Adjusted EBITDA": adj_eb,
                    "Weighted EBITDA": weighted_eb,
                }
            )

        df_eb = pd.DataFrame(rows_eb)

        if use_timing_eb:
            df_eb_display = df_eb[["Year", "EBITDA", "Timing", "Weight (%)", "Adjusted EBITDA", "Weighted EBITDA"]]
        else:
            df_eb_display = df_eb[["Year", "EBITDA", "Weight (%)", "Weighted EBITDA"]]

        df_eb_display = df_eb_display.copy()
        df_eb_display.index = df_eb_display.index + 1
        st.dataframe(format_numeric_columns(df_eb_display), width='stretch')

        maintainable_ebitda = float(df_eb["Weighted EBITDA"].sum())
        st.success(f"Maintainable EBITDA = {maintainable_ebitda:,.2f}")
        S["maintainable_ebitda"] = maintainable_ebitda


# =========================================================
# STEP 4 — MAINTAINABLE EARNINGS (with locked timing)
# =========================================================
st.header("Step 4 — Maintainable Earnings")

dcf_np_all = S.get("dcf_profit_all", None)
if dcf_np_all is None:
    dcf_np_all = S.get("dcf_profit_forecast", {})

# ✅ If no DCF Earnings → SKIP Step 4 (no manual inputs)
if not dcf_np_all:
    st.warning("⚠ No Earnings found from DCF — skipping P/E method.")
    S["maintainable_earnings"] = np.nan

else:
    # ✅ SAFETY: only accept 4-digit year keys (e.g., 2024)
    np_years_all = sorted(
        int(y) for y in dcf_np_all.keys()
        if str(y).strip().isdigit() and len(str(y).strip()) == 4
    )

    if not np_years_all:
        st.warning("⚠ DCF Earnings found, but no valid 4-digit year keys — skipping P/E method.")
        S["maintainable_earnings"] = np.nan

    else:
        np_min_year = min(np_years_all)
        np_max_year = max(np_years_all)

        S.setdefault("comp_np_start_year", np_min_year)
        S.setdefault("comp_np_end_year", np_max_year)
        S.setdefault("comp_np_weights", {})
        S.setdefault("comp_use_timing_np", True)
        # ---------------------------------------------------------
        # AUTO-SYNC Earnings weighting from EBITDA weighting
        # (same years + same weights)
        # ---------------------------------------------------------
        S.setdefault("comp_sync_np_to_eb", True)

        sync_np_to_eb = st.checkbox(
            "Auto-use the SAME years & weights as EBITDA (recommended)",
            value=bool(S.get("comp_sync_np_to_eb", True)),
            key="comp_sync_np_to_eb_checkbox",
        )
        S["comp_sync_np_to_eb"] = bool(sync_np_to_eb)

        if sync_np_to_eb:
            # Copy start/end year from EBITDA section
            eb_sy = int(S.get("comp_eb_start_year", np_min_year))
            eb_ey = int(S.get("comp_eb_end_year", np_max_year))

            # Clamp within NP available year range
            eb_sy = max(eb_sy, np_min_year)
            eb_ey = min(eb_ey, np_max_year)

            S["comp_np_start_year"] = eb_sy
            S["comp_np_end_year"] = eb_ey

            # Copy per-year weights from EBITDA section
            eb_w = S.get("comp_eb_weights", {}) or {}
            S["comp_np_weights"] = {str(y): float(eb_w.get(str(y), 0.0)) for y in range(eb_sy, eb_ey + 1)}

            # IMPORTANT: also prefill the Earnings weight widgets (so UI matches)
            for y in range(eb_sy, eb_ey + 1):
                S[f"comp_np_weight_{y}"] = float(S["comp_np_weights"].get(str(y), 0.0))

            st.info("✅ Earnings years & weights copied from EBITDA automatically.")
        # ---------------------------------------------------------
        # ✅ AUTO-SYNC timing toggle: if EBITDA timing is OFF, Earnings timing must also be OFF
        # ---------------------------------------------------------
        use_timing_eb = bool(S.get("comp_use_timing_eb", True))  # from Step 3

        # If EBITDA timing is OFF, force Earnings timing OFF (also forces UI key)
        if not use_timing_eb:
            S["comp_use_timing_np"] = False
            S["comp_use_timing_np_checkbox"] = False

        use_timing_np = st.checkbox(
            "Apply timing effect from DCF to Earnings?",
            value=bool(S.get("comp_use_timing_np", True)),
            key="comp_use_timing_np_checkbox",
            disabled=(not use_timing_eb),  # lock it when EBITDA timing is OFF
        )
        S["comp_use_timing_np"] = bool(use_timing_np)

        # Show locked years when sync is ON, otherwise allow manual selection
        if sync_np_to_eb:
            np_start_year = int(S["comp_np_start_year"])
            np_end_year = int(S["comp_np_end_year"])

            c_np1, c_np2 = st.columns(2)
            with c_np1:
                st.number_input(
                    "NP Start Year (auto from EBITDA)",
                    value=int(np_start_year),
                    disabled=True,
                    key="np_start_locked",
                )
            with c_np2:
                st.number_input(
                    "NP End Year (auto from EBITDA)",
                    value=int(np_end_year),
                    disabled=True,
                    key="np_end_locked",
                )

        else:
            c_np1, c_np2 = st.columns(2)
            with c_np1:
                np_start_year = st.number_input(
                    "NP Start Year",
                    value=int(S.get("comp_np_start_year", np_min_year)),
                    step=1,
                    key="comp_np_start_year_input"
                )
            with c_np2:
                np_end_year = st.number_input(
                    "NP End Year",
                    value=int(S.get("comp_np_end_year", np_max_year)),
                    step=1,
                    key="comp_np_end_year_input"
                )

            # clamp
            np_start_year = int(max(np_start_year, np_min_year))
            np_end_year = int(min(np_end_year, np_max_year))
            if np_end_year < np_start_year:
                st.error("❌ NP End Year cannot be before Start Year.")
                st.stop()

            S["comp_np_start_year"] = np_start_year
            S["comp_np_end_year"] = np_end_year

        selected_np_years = list(range(np_start_year, np_end_year + 1))
        st.subheader("Earnings Weighting")

        rows_np = []
        base_timing = float(S.get("comp_timing_base", 0.0))

        for idx, yr in enumerate(selected_np_years):
            np_val = float(dcf_np_all.get(str(yr), 0.0))
            default_w = float(S.get(f"comp_np_weight_{yr}", S["comp_np_weights"].get(str(yr), 0.0)))

            if not use_timing_np:
                timing_val = 1.0
            else:
                timing_val = base_timing + idx

            c1, c2, c4 = st.columns([1, 2, 1])
            with c1:
                st.number_input(f"Earnings Year {yr}", value=int(yr), disabled=True, key=f"comp_np_year_display_{yr}")
            with c2:
                st.number_input(f"Earnings {yr}", value=np_val, disabled=True, format="%.2f", key=f"comp_np_value_display_{yr}")
            with c4:
                weight_val = st.number_input(
                    f"NP Weight {yr} (%)",
                    value=float(default_w),
                    step=0.1,
                    format="%.2f",
                    key=f"comp_np_weight_{yr}",
                )
            S["comp_np_weights"][str(yr)] = float(weight_val)
            adj_np = np_val * timing_val
            weighted_np = adj_np * weight_val / 100.0
            rows_np.append(
                {
                    "Year": int(yr),
                    "Earnings": np_val,
                    "Timing": timing_val if use_timing_np else np.nan,
                    "Weight (%)": weight_val,
                    "Adjusted Earnings": adj_np,
                    "Weighted Earnings": weighted_np,
                }
            )
        df_np = pd.DataFrame(rows_np)
        if use_timing_np:
            df_np_display = df_np[["Year", "Earnings", "Timing", "Weight (%)", "Adjusted Earnings", "Weighted Earnings"]]
        else:
            df_np_display = df_np[["Year", "Earnings", "Weight (%)", "Weighted Earnings"]]
        df_np_display = df_np_display.copy()
        df_np_display.index = df_np_display.index + 1
        st.dataframe(format_numeric_columns(df_np_display), width='stretch')
        maintainable_earnings = float(df_np["Weighted Earnings"].sum())
        st.success(f"Maintainable Earnings = {maintainable_earnings:,.2f}")
        S["maintainable_earnings"] = maintainable_earnings
# =========================================================
# STEP 5 — BOOK VALUE & NET DEBT
# =========================================================
st.header("Step 5 — Book Value & Net Debt")
# ✅ Pull Beginning Book Value from Banking page (Totals / BV)
bank_outputs = (S.get("bank", {}) or {}).get("outputs", {}) or {}
bank_book_equity = bank_outputs.get("book_equity_0", None)  # Beginning Book Value (Total Equity)
# If user hasn't typed anything yet, auto-fill book equity from banking
if bank_book_equity is not None:
    # only auto-set if user hasn't created/edited the input widget yet
    if "book_equity_input" not in S:
        S["book_equity"] = float(bank_book_equity)
        S["book_equity_input"] = float(bank_book_equity)

book_equity_default = float(S.get("book_equity", 0.0))
net_debt_default = float(S.get("net_debt", 0.0))

book_equity = st.number_input(
    "Book Equity (USD)",
    value=book_equity_default,
    step=1000.0,
    format="%.2f",   # ⚠ no commas here
    key="book_equity_input"
)
S["book_equity"] = float(book_equity)

# Pretty display with commas (read-only)
st.caption(f"💰 Book Equity: **{book_equity:,.2f} USD**")

net_debt = st.number_input(
    "Net Debt (USD)",
    value=net_debt_default,
    step=1000.0,
    format="%.2f",   # ⚠ no commas here
    key="net_debt_input"
)
S["net_debt"] = float(net_debt)

# Pretty display with commas (read-only)
st.caption(f"💳 Net Debt: **{net_debt:,.2f} USD**")

# =========================================================
# STEP 6 — FINAL EQUITY VALUES
# =========================================================
st.header("Step 6 — Computed Equity Values")

maintainable_ebitda = S.get("maintainable_ebitda", np.nan)
maintainable_earnings = S.get("maintainable_earnings", np.nan)

equity_ev = np.nan
equity_pb = np.nan
equity_pe = np.nan

# EV/EBITDA only if EBITDA exists
if maintainable_ebitda is not None and np.isfinite(float(maintainable_ebitda)) and not pd.isna(implied_ev):
    equity_ev = implied_ev * float(maintainable_ebitda) - net_debt

# P/B works as long as Book Equity exists
if book_equity is not None and np.isfinite(float(book_equity)) and not pd.isna(implied_pb):
    equity_pb = implied_pb * float(book_equity)

# P/E only if Earnings exists
if maintainable_earnings is not None and np.isfinite(float(maintainable_earnings)) and not pd.isna(implied_pe):
    equity_pe = implied_pe * float(maintainable_earnings)


S["value_ev_ebitda"] = float(equity_ev)
S["value_pbv"] = float(equity_pb)
S["value_pe"] = float(equity_pe)

df_res = pd.DataFrame(
    {"Method": ["EV/EBITDA", "P/B", "P/E"], "Equity Value (USD)": [equity_ev, equity_pb, equity_pe]}
)
st.dataframe(format_numeric_columns(df_res), width='stretch')
# =========================================================
# ✅ DOWNLOAD EXCEL (NEAT + FORMULAS) — COMPARABLES EXPORT
# =========================================================
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def _style_range(ws, cell_range, bold=False, fill=None, align_center=False, border=True, font_color=None):
    thin = Side(style="thin", color="000000")
    b = Border(left=thin, right=thin, top=thin, bottom=thin) if border else None
    for row in ws[cell_range]:
        for c in row:
            if bold:
                c.font = Font(bold=True, color=font_color or c.font.color)
            if fill is not None:
                c.fill = fill
            if align_center:
                c.alignment = Alignment(horizontal="center", vertical="center")
            if b is not None:
                c.border = b

def build_comps_excel_with_formulas(S, df_comps) -> bytes:
    wb = Workbook()

    # ---------- Styles ----------
    header_fill = PatternFill("solid", fgColor="0A1B33")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    money_fmt = '#,##0.00'
    pct_fmt = '0.00%'
    mult_fmt = '0.00'

    # ============================
    # Sheet 1: Comps_Input
    # ============================
    ws1 = wb.active
    ws1.title = "Comps_Input"

    ws1["B1"] = "Comparable Company"
    ws1["B1"].font = title_font

    headers = ["Company", "Country", "EV/EBITDA", "P/B", "P/E", "Include_EV", "Include_PB", "Include_PE"]
    start_row = 3
    start_col = 2  # column B

    for j, h in enumerate(headers, start=start_col):
        c = ws1.cell(row=start_row, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Write comps rows
    r = start_row + 1
    for _, row in df_comps.iterrows():
        ws1.cell(r, start_col + 0, row["Company"])
        ws1.cell(r, start_col + 1, "")  # Country (optional)
        ws1.cell(r, start_col + 2, float(row["EV/EBITDA"]) if pd.notna(row["EV/EBITDA"]) else None)
        ws1.cell(r, start_col + 3, float(row["P/B"]) if pd.notna(row["P/B"]) else None)
        ws1.cell(r, start_col + 4, float(row["P/E"]) if pd.notna(row["P/E"]) else None)

        # TRUE/FALSE flags (Excel friendly)
        ws1.cell(r, start_col + 5, bool(row["Include_EV"]))
        ws1.cell(r, start_col + 6, bool(row["Include_PB"]))
        ws1.cell(r, start_col + 7, bool(row["Include_PE"]))

        # formats
        ws1.cell(r, start_col + 2).number_format = mult_fmt
        ws1.cell(r, start_col + 3).number_format = mult_fmt
        ws1.cell(r, start_col + 4).number_format = mult_fmt
        r += 1

    end_row = r - 1

    # Borders + widths
    _style_range(ws1, f"B{start_row}:I{end_row}", border=True)
    for col, w in zip(["B","C","D","E","F","G","H","I"], [30,16,12,10,10,12,12,12]):
        ws1.column_dimensions[col].width = w

    # ============================
    # Sheet 2: Multiples
    # ============================
    ws2 = wb.create_sheet("Multiples")
    ws2["B1"] = "Multiples Summary"
    ws2["B1"].font = title_font

    # Discount cell (user input from session)
    ws2["B3"] = "Discount (%)"
    ws2["C3"] = float(S.get("discount_pct", 25.0)) / 100.0
    ws2["C3"].number_format = pct_fmt
    ws2["B3"].font = bold_font

    # Table headers
    ws2_headers = ["Multiple", "Average", "Implied"]
    for j, h in enumerate(ws2_headers, start=2):
        c = ws2.cell(row=5, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Formulas using AVERAGEIF on Include flags
    # Range in Comps_Input:
    # EV: D, PB: E, PE: F   | flags: G,H,I
    # Data rows: start_row+1 .. end_row
    drow1 = start_row + 1
    drow2 = end_row

    ws2["B6"] = "EV/EBITDA"
    ws2["C6"] = f'=IFERROR(AVERAGEIF(Comps_Input!$G${drow1}:$G${drow2},TRUE,Comps_Input!$D${drow1}:$D${drow2}),"")'
    ws2["D6"] = f'=IF(C6="","",C6*(1-$C$3))'

    ws2["B7"] = "P/B"
    ws2["C7"] = f'=IFERROR(AVERAGEIF(Comps_Input!$H${drow1}:$H${drow2},TRUE,Comps_Input!$E${drow1}:$E${drow2}),"")'
    ws2["D7"] = f'=IF(C7="","",C7*(1-$C$3))'

    ws2["B8"] = "P/E"
    ws2["C8"] = f'=IFERROR(AVERAGEIF(Comps_Input!$I${drow1}:$I${drow2},TRUE,Comps_Input!$F${drow1}:$F${drow2}),"")'
    ws2["D8"] = f'=IF(C8="","",C8*(1-$C$3))'

    for rr in [6,7,8]:
        ws2[f"C{rr}"].number_format = mult_fmt
        ws2[f"D{rr}"].number_format = mult_fmt

    _style_range(ws2, "B5:D8", border=True)
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14

    # ============================
    # Sheet 3: EBITDA_Maintainable
    # ============================
    ws3 = wb.create_sheet("EBITDA_Maintainable")
    ws3["B1"] = "Maintainable EBITDA (with timing + weights)"
    ws3["B1"].font = title_font

    ws3["B3"] = "Use Timing?"
    ws3["C3"] = bool(S.get("comp_use_timing_eb", True))
    ws3["B4"] = "Base Timing"
    ws3["C4"] = float(S.get("comp_timing_base", 1.0))

    ws3["B3"].font = bold_font
    ws3["B4"].font = bold_font

    # Pull years + EBITDA from session
    dcf_eb_all = S.get("dcf_ebitda_all", None) or S.get("dcf_ebitda_forecast", {}) or {}
    eb_sy = int(S.get("comp_eb_start_year", 0) or 0)
    eb_ey = int(S.get("comp_eb_end_year", 0) or 0)
    eb_years = list(range(eb_sy, eb_ey + 1)) if eb_sy and eb_ey and eb_ey >= eb_sy else []

    # Table
    headers3 = ["Year", "EBITDA", "Timing", "Weight (%)", "Adjusted EBITDA", "Weighted EBITDA"]
    for j, h in enumerate(headers3, start=2):
        c = ws3.cell(row=6, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    r0 = 7
    for idx, yr in enumerate(eb_years):
        ws3.cell(r0+idx, 2, yr)
        ws3.cell(r0+idx, 3, float(dcf_eb_all.get(str(yr), 0.0)))

        # Timing formula:
        # =IF($C$3, $C$4 + (ROW()-7), 1)
        ws3.cell(r0+idx, 4, f'=IF($C$3,$C$4+{idx},1)')

        # Weight from session (store as percent)
        w = float((S.get("comp_eb_weights", {}) or {}).get(str(yr), 0.0))
        ws3.cell(r0+idx, 5, w/100.0)

        # Adjusted EBITDA = EBITDA * Timing
        ws3.cell(r0+idx, 6, f"=C{r0+idx}*D{r0+idx}")
        # Weighted EBITDA = Adjusted * Weight
        ws3.cell(r0+idx, 7, f"=F{r0+idx}*E{r0+idx}")

        ws3.cell(r0+idx, 3).number_format = money_fmt
        ws3.cell(r0+idx, 4).number_format = '0.0000'
        ws3.cell(r0+idx, 5).number_format = pct_fmt
        ws3.cell(r0+idx, 6).number_format = money_fmt
        ws3.cell(r0+idx, 7).number_format = money_fmt

    last = r0 + len(eb_years) - 1 if eb_years else 7

    # Total maintainable EBITDA
    ws3["B" + str(last+2)] = "Maintainable EBITDA"
    ws3["B" + str(last+2)].font = bold_font
    ws3["G" + str(last+2)] = f"=SUM(G{r0}:G{last})"
    ws3["G" + str(last+2)].font = bold_font
    ws3["G" + str(last+2)].number_format = money_fmt

    _style_range(ws3, f"B6:G{last}", border=True)
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 12
    ws3.column_dimensions["F"].width = 18
    ws3.column_dimensions["G"].width = 18

    # ============================
    # Sheet 4: Earnings_Maintainable
    # ============================
    ws4 = wb.create_sheet("Earnings_Maintainable")
    ws4["B1"] = "Maintainable Earnings (with timing + weights)"
    ws4["B1"].font = title_font

    ws4["B3"] = "Use Timing?"
    ws4["C3"] = bool(S.get("comp_use_timing_np", True))
    ws4["B4"] = "Base Timing"
    ws4["C4"] = float(S.get("comp_timing_base", 1.0))
    ws4["B3"].font = bold_font
    ws4["B4"].font = bold_font

    dcf_np_all = S.get("dcf_profit_all", None) or S.get("dcf_profit_forecast", {}) or {}
    np_sy = int(S.get("comp_np_start_year", 0) or 0)
    np_ey = int(S.get("comp_np_end_year", 0) or 0)
    np_years = list(range(np_sy, np_ey + 1)) if np_sy and np_ey and np_ey >= np_sy else []

    headers4 = ["Year", "Earnings", "Timing", "Weight (%)", "Adjusted Earnings", "Weighted Earnings"]
    for j, h in enumerate(headers4, start=2):
        c = ws4.cell(row=6, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    r0 = 7
    for idx, yr in enumerate(np_years):
        ws4.cell(r0+idx, 2, yr)
        ws4.cell(r0+idx, 3, float(dcf_np_all.get(str(yr), 0.0)))

        ws4.cell(r0+idx, 4, f'=IF($C$3,$C$4+{idx},1)')

        w = float((S.get("comp_np_weights", {}) or {}).get(str(yr), 0.0))
        ws4.cell(r0+idx, 5, w/100.0)

        ws4.cell(r0+idx, 6, f"=C{r0+idx}*D{r0+idx}")
        ws4.cell(r0+idx, 7, f"=F{r0+idx}*E{r0+idx}")

        ws4.cell(r0+idx, 3).number_format = money_fmt
        ws4.cell(r0+idx, 4).number_format = '0.0000'
        ws4.cell(r0+idx, 5).number_format = pct_fmt
        ws4.cell(r0+idx, 6).number_format = money_fmt
        ws4.cell(r0+idx, 7).number_format = money_fmt

    last = r0 + len(np_years) - 1 if np_years else 7

    ws4["B" + str(last+2)] = "Maintainable Earnings"
    ws4["B" + str(last+2)].font = bold_font
    ws4["G" + str(last+2)] = f"=SUM(G{r0}:G{last})"
    ws4["G" + str(last+2)].font = bold_font
    ws4["G" + str(last+2)].number_format = money_fmt

    _style_range(ws4, f"B6:G{last}", border=True)
    for col, w in zip(["B","C","D","E","F","G"], [10,18,12,12,18,18]):
        ws4.column_dimensions[col].width = w

    # ============================
    # Sheet 5: Equity_Values
    # ============================
    ws5 = wb.create_sheet("Equity_Values")
    ws5["B1"] = "Computed Equity Values"
    ws5["B1"].font = title_font

    # Inputs needed (from your Step 5)
    ws5["B3"] = "Book Equity"
    ws5["C3"] = float(S.get("book_equity", 0.0))
    ws5["B4"] = "Net Debt"
    ws5["C4"] = float(S.get("net_debt", 0.0))
    ws5["B3"].font = bold_font
    ws5["B4"].font = bold_font
    ws5["C3"].number_format = money_fmt
    ws5["C4"].number_format = money_fmt

    # Link maintainables
    ws5["B6"] = "Maintainable EBITDA"
    ws5["C6"] = "=EBITDA_Maintainable!G" + str((ws3.max_row))  # last maintainable cell
    ws5["B7"] = "Maintainable Earnings"
    ws5["C7"] = "=Earnings_Maintainable!G" + str((ws4.max_row))  # last maintainable cell

    ws5["C6"].number_format = money_fmt
    ws5["C7"].number_format = money_fmt
    ws5["B6"].font = bold_font
    ws5["B7"].font = bold_font

    # Equity table
    ws5_headers = ["Method", "Equity Value (USD)"]
    for j, h in enumerate(ws5_headers, start=2):
        c = ws5.cell(row=9, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Implied multiples link from Multiples sheet:
    # EV implied at D6, PB implied at D7, PE implied at D8
    ws5["B10"] = "EV/EBITDA"
    ws5["C10"] = "=IF(Multiples!D6=\"\",\"\",Multiples!D6*$C$6-$C$4)"

    ws5["B11"] = "P/B"
    ws5["C11"] = "=IF(Multiples!D7=\"\",\"\",Multiples!D7*$C$3)"

    ws5["B12"] = "P/E"
    ws5["C12"] = "=IF(Multiples!D8=\"\",\"\",Multiples!D8*$C$7)"

    for rr in [10,11,12]:
        ws5[f"C{rr}"].number_format = money_fmt

    _style_range(ws5, "B9:C12", border=True)
    ws5.column_dimensions["B"].width = 18
    ws5.column_dimensions["C"].width = 22

    # Save
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

excel_bytes = build_comps_excel_with_formulas(S, df_comps)

st.download_button(
    label="⬇️ Download Comparables (Excel with formulas)",
    data=excel_bytes,
    file_name="comparables_with_formulas.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
